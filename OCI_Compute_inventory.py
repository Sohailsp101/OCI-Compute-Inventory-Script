import oci
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, PatternFill
)
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone
import sys


# ---------------------------------------------------------------------------
CONFIG_PROFILE = "DEFAULT"
OUTPUT_FILE = f"OCI_Compute_Report_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

SHEET_ORDER = ["Summary", "PROD", "DEV", "UAT", "DR", "OCVS"]


COLUMNS = [
    "S.No", "Instance Name", "Instance OCID", "Private IP", "Public IP",
    "Compartment", "Subnet", "OCPU", "Memory (GB)", "State", "Shape",
    "Operating System", "Environment Tag", "Application Tag", "Creation Time",
]


STATE_COLORS = {
    "RUNNING":      "FF92D050", 
    "STOPPED":      "FFFF0000", 
    "STOPPING":     "FFED7D31",  
    "STARTING":     "FFFFC000",  
    "TERMINATED":   "FF808080", 
    "PROVISIONING": "FF4472C4",
}

HEADER_FILL  = PatternFill("solid", start_color="FF1F497D") 
ALT_ROW_FILL = PatternFill("solid", start_color="FFF2F2F2")   
WHITE_FILL   = PatternFill("solid", start_color="FFFFFFFF")

HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFFFF", size=10)
DATA_FONT    = Font(name="Arial", size=10)
BOLD_FONT    = Font(name="Arial", bold=True, size=10)

THIN_BORDER = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
WRAP   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------

def get_config():
    return oci.config.from_file(profile_name=CONFIG_PROFILE)


def get_all_compartments(identity_client, tenancy_id):
    """Return all active compartments including root."""
    compartments = [
        type("C", (), {"id": tenancy_id, "name": "root"})()
    ]
    paginator = identity_client.list_compartments(
        tenancy_id,
        compartment_id_in_subtree=True,
        lifecycle_state="ACTIVE",
        limit=1000,
    )
    compartments.extend(paginator.data)
    return compartments


def get_subscribed_regions(identity_client, tenancy_id):
    """Return list of subscribed region names."""
    resp = identity_client.list_region_subscriptions(tenancy_id)
    return [r.region_name for r in resp.data]


def get_vnic_details(compute_client, network_client, instance_id, compartment_id):
    """Return (private_ip, public_ip, subnet_name) for the primary VNIC."""
    try:
        attachments = compute_client.list_vnic_attachments(
            compartment_id=compartment_id,
            instance_id=instance_id,
        ).data
        if not attachments:
            return "N/A", "N/A", "N/A"
        vnic = network_client.get_vnic(attachments[0].vnic_id).data
        subnet_name = "N/A"
        try:
            subnet = network_client.get_subnet(vnic.subnet_id).data
            subnet_name = subnet.display_name or "N/A"
        except Exception:
            pass
        return (
            vnic.private_ip or "N/A",
            vnic.public_ip or "N/A",
            subnet_name,
        )
    except Exception:
        return "N/A", "N/A", "N/A"


def get_image_os(compute_client, image_id):
    """Return OS string from image."""
    try:
        img = compute_client.get_image(image_id).data
        os_name = img.operating_system or ""
        os_ver  = img.operating_system_version or ""
        return f"{os_name} {os_ver}".strip() or "N/A"
    except Exception:
        return "N/A"


def classify_by_name(instance_name):
    """Classify instance into a sheet based on keywords in its name (case-insensitive).
    Priority order: DR → OCVS → UAT → DEV → PROD (fallback).
    """
    name = (instance_name or "").upper()
    if "DR" in name:
        return "DR"
    if "OCVS" in name:
        return "OCVS"
    if "UAT" in name:
        return "UAT"
    if "DEV" in name:
        return "DEV"
    return "PROD"


# ---------------------------------------------------------------------------

def collect_instances():
    print("Initializing OCI config …")
    config = get_config()
    tenancy_id = config["tenancy"]

    identity_client = oci.identity.IdentityClient(config)

    print("Fetching subscribed regions …")
    regions = get_subscribed_regions(identity_client, tenancy_id)
    print(f"  Subscribed regions: {regions}")

    print("Fetching all compartments …")
    compartments = get_all_compartments(identity_client, tenancy_id)
    print(f"  Total compartments: {len(compartments)}")

    comp_map = {c.id: c.name for c in compartments}

    all_instances = []

    for region in regions:
        print(f"\n[Region: {region}]")
        region_config = dict(config)
        region_config["region"] = region

        compute_client  = oci.core.ComputeClient(region_config)
        network_client  = oci.core.VirtualNetworkClient(region_config)

        for comp in compartments:
            try:
                instances = oci.pagination.list_call_get_all_results(
                    compute_client.list_instances,
                    compartment_id=comp.id,
                ).data
            except Exception as e:

                continue

            active = [i for i in instances if i.lifecycle_state != "TERMINATED"]
            if active:
                print(f"  {comp.name}: {len(active)} instance(s)")

            for inst in active:
                priv_ip, pub_ip, subnet = get_vnic_details(
                    compute_client, network_client, inst.id, comp.id
                )

                freeform = inst.freeform_tags or {}
                defined  = inst.defined_tags or {}

                # Resolve Environment tag (freeform first, then defined)
                env_tag = (
                    freeform.get("Environment")
                    or freeform.get("environment")
                    or next(
                        (defined[ns].get("Environment") or defined[ns].get("environment")
                         for ns in defined if "Environment" in defined[ns] or "environment" in defined[ns]),
                        None,
                    )
                )

                
                app_tag = (
                    freeform.get("Application")
                    or freeform.get("application")
                    or next(
                        (defined[ns].get("Application") or defined[ns].get("application")
                         for ns in defined if "Application" in defined[ns] or "application" in defined[ns]),
                        None,
                    )
                )

                shape_cfg = inst.shape_config
                ocpus  = shape_cfg.ocpus  if shape_cfg else None
                memory = shape_cfg.memory_in_gbs if shape_cfg else None

                os_name = get_image_os(compute_client, inst.image_id) if inst.image_id else "N/A"

                creation_time = inst.time_created
                if creation_time:
                   
                    creation_time = creation_time.strftime("%Y-%m-%d %H:%M:%S UTC")

                sheet = classify_by_name(inst.display_name)

                all_instances.append({
                    "Instance Name":    inst.display_name or "N/A",
                    "Instance OCID":    inst.id,
                    "Private IP":       priv_ip,
                    "Public IP":        pub_ip,
                    "Compartment":      comp_map.get(inst.compartment_id, inst.compartment_id),
                    "Subnet":           subnet,
                    "OCPU":             ocpus if ocpus is not None else "N/A",
                    "Memory (GB)":      memory if memory is not None else "N/A",
                    "State":            inst.lifecycle_state,
                    "Shape":            inst.shape or "N/A",
                    "Operating System": os_name,
                    "Environment Tag":  env_tag or "N/A",
                    "Application Tag":  app_tag or "N/A",
                    "Creation Time":    creation_time or "N/A",
                    "_sheet":           sheet,
                    "_region":          region,
                })

    print(f"\nTotal instances collected: {len(all_instances)}")
    return all_instances


# ---------------------------------------------------------------------------

def apply_header(ws, columns):
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font       = HEADER_FONT
        cell.fill       = HEADER_FILL
        cell.alignment  = CENTER
        cell.border     = THIN_BORDER


def apply_row(ws, row_num, values, state=None):
    fill = ALT_ROW_FILL if row_num % 2 == 0 else WHITE_FILL
    numeric_cols = {8, 9} 

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font   = DATA_FONT
        cell.border = THIN_BORDER

        if col_idx == 10 and state and state in STATE_COLORS: 
            cell.fill      = PatternFill("solid", start_color=STATE_COLORS[state])
            cell.font      = Font(name="Arial", size=10, bold=True,
                                  color="FFFFFFFF" if state not in ("STARTING",) else "FF000000")
            cell.alignment = CENTER
        elif col_idx == 3:  
            cell.fill      = fill
            cell.alignment = WRAP
        elif col_idx in (8, 9): 
            cell.fill      = fill
            cell.alignment = CENTER
        else:
            cell.fill      = fill
            cell.alignment = LEFT


def autofit_columns(ws, columns):
    for col_idx, col_name in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(col_name)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    max_len = max(max_len, cell_len)
                except Exception:
                    pass
        if col_name == "Instance OCID":
            ws.column_dimensions[col_letter].width = 40
        elif col_name in ("Instance Name", "Compartment", "Operating System", "Subnet"):
            ws.column_dimensions[col_letter].width = min(max_len + 2, 35)
        else:
            ws.column_dimensions[col_letter].width = min(max_len + 2, 25)


def build_data_sheet(wb, sheet_name, rows):
    ws = wb[sheet_name]
    apply_header(ws, COLUMNS)

    rows_sorted = sorted(rows, key=lambda x: (x["Compartment"], x["Instance Name"]))

    for i, inst in enumerate(rows_sorted, 2):
        s_no = i - 1
        values = [
            s_no,
            inst["Instance Name"],
            inst["Instance OCID"],
            inst["Private IP"],
            inst["Public IP"],
            inst["Compartment"],
            inst["Subnet"],
            inst["OCPU"],
            inst["Memory (GB)"],
            inst["State"],
            inst["Shape"],
            inst["Operating System"],
            inst["Environment Tag"],
            inst["Application Tag"],
            inst["Creation Time"],
        ]
        apply_row(ws, i, values, state=inst["State"])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autofit_columns(ws, COLUMNS)


def build_summary_sheet(wb, all_instances):
    ws = wb["Summary"]

    ws.merge_cells("A1:C1")
    title_cell = ws["A1"]
    title_cell.value     = "OCI Compute Instances — Summary Report"
    title_cell.font      = Font(name="Arial", bold=True, size=14, color="FFFFFFFF")
    title_cell.fill      = HEADER_FILL
    title_cell.alignment = CENTER

    ws.merge_cells("A2:C2")
    ws["A2"].value     = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font      = Font(name="Arial", italic=True, size=10)
    ws["A2"].alignment = CENTER

    ws["A4"].value = "State"
    ws["B4"].value = "Count"
    for cell in (ws["A4"], ws["B4"]):
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = THIN_BORDER

    states = ["RUNNING", "STOPPED", "STOPPING", "STARTING", "PROVISIONING", "TERMINATED"]
    state_counts = {s: sum(1 for i in all_instances if i["State"] == s) for s in states}
    total = len(all_instances)

    state_rows = [("Total Compute Instances", total)] + [(s, state_counts.get(s, 0)) for s in states]
    for r_offset, (label, count) in enumerate(state_rows, 5):
        fill = ALT_ROW_FILL if r_offset % 2 == 0 else WHITE_FILL
        a = ws.cell(row=r_offset, column=1, value=label)
        b = ws.cell(row=r_offset, column=2, value=count)
        for cell in (a, b):
            cell.font      = BOLD_FONT if label == "Total Compute Instances" else DATA_FONT
            cell.fill      = fill
            cell.alignment = CENTER
            cell.border    = THIN_BORDER

    row_start = 5 + len(state_rows) + 2

    ws.cell(row=row_start, column=1).value = "Environment"
    ws.cell(row=row_start, column=2).value = "Count"
    for col in (1, 2):
        cell = ws.cell(row=row_start, column=col)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = THIN_BORDER

    for r_offset, sheet_name in enumerate(["PROD", "DEV", "UAT", "DR", "OCVS"], row_start + 1):
        count = sum(1 for i in all_instances if i["_sheet"] == sheet_name)
        fill  = ALT_ROW_FILL if r_offset % 2 == 0 else WHITE_FILL
        a = ws.cell(row=r_offset, column=1, value=sheet_name)
        b = ws.cell(row=r_offset, column=2, value=count)
        for cell in (a, b):
            cell.font      = DATA_FONT
            cell.fill      = fill
            cell.alignment = CENTER
            cell.border    = THIN_BORDER

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15


# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("OCI Multi-Region Compute Report Generator")
    print("=" * 60)

    all_instances = collect_instances()

    if not all_instances:
        print("\nNo instances found. Exiting.")
        sys.exit(0)

    sheet_data = {s: [] for s in SHEET_ORDER if s != "Summary"}
    for inst in all_instances:
        sheet_data[inst["_sheet"]].append(inst)

    print("\nBuilding Excel workbook …")
    wb = openpyxl.Workbook()
    wb.remove(wb.active) 

    for sheet_name in SHEET_ORDER:
        wb.create_sheet(title=sheet_name)

    build_summary_sheet(wb, all_instances)

    for sheet_name in SHEET_ORDER:
        if sheet_name == "Summary":
            continue
        rows = sheet_data.get(sheet_name, [])
        print(f"  {sheet_name}: {len(rows)} instance(s)")
        build_data_sheet(wb, sheet_name, rows)

    wb.save(OUTPUT_FILE)
    print(f"\nReport saved: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
